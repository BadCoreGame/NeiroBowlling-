# библиотеки локальные -----------------------------
import os # работа с файлами и директориями
import shutil # работа с файлами и директориями
from pathlib import Path # работа с путями
from datetime import datetime # работа с датой и временем
import tkinter as tk #для создания окна
from tkinter import filedialog, messagebox # tkinter импорт
import customtkinter as ctk #tkinter модуль красоты
from PIL import Image, ImageTk #tkinter оьраьотка для него изображений
import pandas as pd #для создания xlsx файла
import openpyxl as xl #для чтения и изменения xlsx файла
# библиотеки для нейронки
import cv2
import numpy as np
import matplotlib.pyplot as plt
import easyocr
import re
# оптимизация


# внешние файлы -----------------------------
class outside():
    path_icno = ".\\app\\icon.png" #путь к иконке окна
    path_logo = ".\\app\\logo.png" #путь к иконке окна
    path_xlsx = ".\\app\\pattern.xlsx" #путь к xlsx шаблону


# общие переменные --------------------------
class peremens():# глобальные переменные
    Path_dir=""  #путь к папке
    loaded_images_paths = [] # загруженные пути к картинкам
    filepath="A:\\Work\\NeiroBowlling\\icon.png" #путь к фотке
    game_round=0 #игра за день
    buffer_path = ""# для работы глобального xlsx файла
    destination_array = []
    
    # вывод с нейронки (заготовки для проверки без ии)
    players=["player1","player2","player3","player4","player5"] #игроки
    scores=[] #баллы


# логирование
def log(name): # логирование
    timestamp = datetime.now().strftime('%H:%M:%S')
    print(f"{timestamp}-{name}" )
    names = f"{timestamp}-{name}\n"
    window.dev_log.configure(state="normal") 
    window.dev_log.insert(tk.END,names) # логирование
    window.dev_log.configure(state="disabled")

#  нейронка ---------------------------------

def show_image(title, image, cmap='gray'):
    #Отображение изображения с использованием matplotlib.
    # plt.figure(figsize=(10, 10))
    # plt.title(title)
    # plt.imshow(image, cmap=cmap)
    # plt.axis('off')
    # plt.show()
    print("ok")

def resize_image(image, max_width=800, max_height=600):
    # Получаем высоту и ширину изображения
    h, w = image.shape[:2]
    # Вычисляем коэффициент масштабирования, чтобы сохранить пропорции
    scale = min(max_width / w, max_height / h)
    # Если изображение больше максимальных размеров, изменяем его размер
    if scale < 1.0:
        # Изменяем размер изображения с использованием интерполяции
        resized = cv2.resize(image, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)
        return resized, scale  # Возвращаем измененное изображение и коэффициент масштабирования
    show_image("fdfds", resized)
    return image, 1.0  # Если изображение меньше максимальных размеров, возвращаем его без изменений

def warp_perspective(image, points):
    width = int(max(
        np.linalg.norm(points[1] - points[0]),
        np.linalg.norm(points[3] - points[2])
    ))
    height = int(max(
        np.linalg.norm(points[3] - points[0]),
        np.linalg.norm(points[2] - points[1])
    ))

    dst = np.array([
        [0, 0],
        [width - 1, 0],
        [width - 1, height - 1],
        [0, height - 1]
    ], dtype="float32")

    matrix = cv2.getPerspectiveTransform(points, dst)
    warped_image = cv2.warpPerspective(image, matrix, (width, height))
    show_image("Изображение с перспективой", warped_image)
    return warped_image

def select_points(image, num_points, max_width=800, max_height=600):
    resized_image, scale = resize_image(image, max_width, max_height)
    points = []

    def click_event(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            points.append((x, y))
            cv2.circle(temp_image, (x, y), 5, (0, 255, 0), -1)
            cv2.imshow("Выбор точек", temp_image)
            if len(points) == num_points:
                cv2.destroyAllWindows()

    temp_image = resized_image.copy()
    cv2.imshow("Выбор точек", temp_image)
    cv2.setMouseCallback("Выбор точек", click_event)
    print(f"Пожалуйста, выберите {num_points} точку(и) на изображении.")
    cv2.waitKey(0)

    if len(points) != num_points:
        raise ValueError(f"Необходимо выбрать ровно {num_points} точек.")
    return np.array(points, dtype="float32") / scale

def get_extreme_boxes(results):
    if not results or len(results) < 4:
        print("Недостаточно данных для анализа.")
        return [], []

    # Сортируем коробки по X-координате верхнего левого угла
    results_sorted_by_left = sorted(results, key=lambda x: x[0][0][0])
    # Сортируем коробки по X-координате нижнего правого угла
    results_sorted_by_right = sorted(results, key=lambda x: x[0][2][0], reverse=True)

    # Возвращаем четыре левых и четыре правых коробки
    left_boxes = results_sorted_by_left[:5]
    right_boxes = results_sorted_by_right[:5]

    return left_boxes, right_boxes

def process_image(image_path):
    # Загрузка изображения
    image = cv2.imread(image_path)
    if image is None:
        print("Ошибка загрузки изображения.")
        return

    # Преобразование в серый цвет
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Бинаризация
    _, binary = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)

    points = select_points(image, 4)
    roi = warp_perspective(binary, points)

    # Распознавание текста
    results = extract_text_from_image(roi)

    if not results:
        print("Текст не обнаружен.")
        return

    # Определение крайних коробок
    left_boxes, right_boxes = get_extreme_boxes(results)

    # Массивы для хранения текста из левых и правых коробок
    # left_texts = []
    right_texts = []

    error_map = {
        "In": "m",
        "0": "o",
        "O": "o",
        "o": "o",
        "5": "s",
        "1": "i",
        "6": "d",
        "D": "d",
        "ol": "d"
    }

    # Обработка левых коробок (сортировка по y)
    # left_boxes_sorted = sorted(left_boxes, key=lambda box: box[0][0][1])

    # for box in left_boxes_sorted:
    #     bbox = box[0]
    #     text = box[1]
    #     # Оставляем только буквы
    #     filtered_text = re.sub(r'[^a-zA-Z]', '', text)

    #     # Проверяем длину текста
    #     if 1 <= len(filtered_text) <= 2:
    #         # Заменяем текст по словарю ошибок, если он есть в error_map
    #         if len(filtered_text) == 2 and filtered_text in error_map:
    #             filtered_text = error_map[filtered_text]

    #         # Добавляем в список
    #         left_texts.append(filtered_text)

    #         # Отрисовка на изображении
    #         l_bbox = bbox[0]
    #         r_bbox = bbox[2]
    #         cv2.rectangle(roi, (int(l_bbox[0]), int(l_bbox[1])), (int(r_bbox[0]), int(r_bbox[1])), (0, 255, 0), 2)
    #         cv2.putText(roi, filtered_text, (int(l_bbox[0]), int(l_bbox[1] - 10)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)

    # Обработка правых коробок (сортировка по y)
    right_boxes_sorted = sorted(right_boxes, key=lambda box: box[0][0][1])

    for box in right_boxes_sorted:
        bbox = box[0]
        text = box[1]
        # Проверяем, состоит ли текст только из цифр
        if text.isdigit() and len(text) <= 3:  # Условие: текст содержит только цифры и до 3 символов
            num_value = int(text)
            if num_value <= 400:  # Убираем значения выше 400
                right_texts.append(num_value)  # Добавляем как число
                l_bbox = bbox[0]
                r_bbox = bbox[2]
                cv2.rectangle(roi, (int(l_bbox[0]), int(l_bbox[1])), (int(r_bbox[0]), int(r_bbox[1])), (255, 0, 0), 2)
                cv2.putText(roi, text, (int(l_bbox[0]), int(l_bbox[1] - 10)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 1)

    # Отображение изображения с коробками
    #show_image("Изображение с крайними коробками", roi)

    # Вывод массивов с текстом
    # print("Текст из левых коробок:", left_texts)
    print("Текст из правых коробок:", right_texts)

    # peremens.players=left_texts
    peremens.scores=right_texts
    log("process finished")

    return right_texts
    # return left_texts, right_texts

def extract_text_from_image(image):
    reader = easyocr.Reader(['en'], gpu=False)  # Используем английский для точности
    # Распознаем текст на всём изображении
    results = reader.readtext(image, allowlist="ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789/X")
    return results

# Локальные функции
class Functions():# функции
    def look(): # функция быстрого просмотра xlsx файлов
        log(f"look directory {peremens.Path_dir}\n")
        if peremens.Path_dir: # проверка на пустоту пути
            peremens.loaded_images_paths = [] #очистка списка
            # дальше поиск файлов по директориям и запись полного пути
            for root, dirs, files in os.walk(peremens.Path_dir):
                for file in files:
                    if file.lower().endswith((".xlsx")):
                        return True
        return False
    
    def clear_xlsm_down(): # загрузка папки
        log(f"cleary directory {peremens.Path_dir}\n")
        if peremens.Path_dir: # проверка на пустоту пути
            peremens.loaded_images_paths = [] #очистка списка
            # дальше поиск файлов по директориям и запись полного пути
            for root, dirs, files in os.walk(peremens.Path_dir):
                for file in files:
                    if file.lower().endswith((".xlsx")):
                        os.remove(os.path.join(root, file))

    def download_images(): # загрузка папки
        peremens.Path_dir= filedialog.askdirectory(title="Выберите папку с изображениями.") # выбор папки
        log(f"download directory {peremens.Path_dir}\n")
        if peremens.Path_dir: # проверка на пустоту пути
            peremens.loaded_images_paths = [] #очистка списка
            # дальше поиск файлов по директориям и запись полного пути
            for root, dirs, files in os.walk(peremens.Path_dir):
                for file in files:
                    if file.lower().endswith((".jpg", ".jpeg", ".png", ".ico", ".bmp", ".tif", ".pcx")):
                        peremens.loaded_images_paths.append(os.path.join(root, file))
            log(f"download images: {peremens.loaded_images_paths[0]}, ...\n")
            print(f"download images: {peremens.loaded_images_paths}")
    
    def save_resultat(): # сохранение результата
        
        def create_path(filepath): # переделка пути для файла resultats.xlsx
            filepath = os.path.dirname(filepath)  # Получаем имя файла из пути
            filepath = os.path.join(filepath, "resultats.xlsx")  # делаем путь до файла
            log(filepath) # логирование
            return filepath # возвращаем путь
        filepath_xslx= create_path(peremens.filepath)
        # это конечно костыль, ну а чего ожидать, для ускорения работы :)

        if peremens.buffer_path == "":  ##запуск запись глобальной
            peremens.buffer_path = filepath_xslx
        elif filepath_xslx != peremens.buffer_path: 
            Functions.save_resultat_global()
            peremens.buffer_path = filepath_xslx
            log(f"File {filepath_xslx} already exist\n")

        if not os.path.exists(filepath_xslx): # проверка на существование файла
            peremens.game_round = 0
            # копируем файл patter.xlsx в папку и переименовываем его в resultats.xlsx
            shutil.copyfile(outside.path_xlsx, filepath_xslx) # копирование файла
            log(f"Create file {filepath_xslx}\n") 
        
        def record_xslx(filepath_xslx): # запись
            # Load in the workbook
            exel = xl.load_workbook(filepath_xslx)
            # поиск людей и добавление им данных, если они есть на страничке "Подробно", иначе дополняем первый столбик именем и добавляем данные
            # не трожьте, произведение искусства из костылей
            peremens.destination_array = peremens.players[:len(peremens.scores)]
            def seek_xslx(exel):
                def balls(player, sas):
                        dsa = peremens.scores[peremens.players.index(player)]
                        if dsa == 100:
                            opa.cell(row=sas, column=(peremens.game_round) * 2 +3).value = 0
                        elif dsa < 100:
                            opa.cell(row=sas, column=(peremens.game_round) * 2 +3).value = -(((100 - peremens.scores[peremens.players.index(player)])//31)+1)
                        elif dsa > 100:
                            opa.cell(row=sas, column=(peremens.game_round) * 2 +3).value = (((peremens.scores[peremens.players.index(player)]- 100)//31)+1)                   
                opa=exel['Подробно']

                
                for player in peremens.destination_array: # поиск людей и добавление им данных, если они есть на страничке "Подробно"
                    for row in opa.rows:# поиск людей и добавление им данных, методом перебора ячеек справа до пустой
                        if row[0].value == player: # если игрок есть, то добавляем данные в правую свободную ячейку
                            opa.cell(row=row[0].row, column=(peremens.game_round + 1) * 2).value = peremens.scores[peremens.destination_array.index(player)]
                            balls(player, row[0].row)
                            break
                    else: # добавляем игрока в конец первого столбца и записываем данные
                        if opa.max_row == 1:
                            opa.cell(row=3, column=1).value = player
                            opa.cell(row=3, column=2).value = peremens.scores[peremens.destination_array.index(player)]
                            balls(player, 3)
                        else:
                            opa.cell(row=opa.max_row + 1, column=1).value = player
                            opa.cell(row=opa.max_row, column=2).value = peremens.scores[peremens.destination_array.index(player)]
                            balls(player, opa.max_row)
            seek_xslx(exel)

            if peremens.game_round < ((exel['Подробно'].max_column-1) // 2 ):
                # добавление строки с данными и index
                peremens.game_round+=1
                exel['Подробно'].cell(row=1, column=peremens.game_round * 2).value = f"Игра {peremens.game_round}"
                exel['Подробно'].cell(row=2, column=peremens.game_round * 2).value = "Результат"
                exel['Подробно'].cell(row=2, column=peremens.game_round * 2 + 1).value = "Баллы"
                exel['Подробно'].merge_cells(start_row=1, start_column=peremens.game_round * 2, end_row=1, end_column=peremens.game_round * 2 + 1)
            # запись в файл
            exel.save(filepath_xslx)
        record_xslx(filepath_xslx)
        log(f"Save in {filepath_xslx}\n")

    def save_resultat_global(): # сохранение глобального результата
        filepath_core = os.path.join(peremens.Path_dir, "global_resultats.xlsx")
        try:
            rezultat_global = xl.load_workbook(filepath_core)
        except FileNotFoundError:
            rezultat_global = xl.Workbook(filepath_core)
            log(f"Create file {filepath_core}\n") 
        
        rezultat_single = xl.load_workbook(peremens.buffer_path)

        # Открываем исходный файл
    
        # Проверяем, существует ли лист с указанным именем
        if "Подробно" not in rezultat_single.sheetnames:
            print(f"Лист 'Подробно' не найден в файле '{rezultat_single}'.")
            return

        rezultat_single = rezultat_single["Подробно"]

        def perevod(path, base_path):
            # Приводим пути к стандартному виду
            normalized_path = os.path.normpath(path)
            normalized_base_path = os.path.normpath(base_path)

            # Получаем абсолютные пути
            absolute_path = Path(normalized_path).absolute()
            absolute_base_path = Path(normalized_base_path).absolute()

            # Получаем относительный путь к родительской директории
            relative_directory = absolute_path.parent.relative_to(absolute_base_path)

            # Заменяем знак '\' на '-'
            return str(relative_directory).replace('\\', '_')

        # Копируем лист
        pathq = perevod(peremens.buffer_path, peremens.Path_dir)
        print(f"Копируем лист '{pathq}'...")
        rezultat_global_2 = rezultat_global.create_sheet(pathq)

        for row in rezultat_single.iter_rows(values_only=True):
            rezultat_global_2.append(row)

        # Сохраняем целевой файл
        rezultat_global.save(filepath_core)
        print(f"Лист 'Подробно' успешно скопирован в файл '{rezultat_global}'.")

    def start():# старт обработки фотографий
        # глобальные переменные
        window.tabview.set("devtool")
        if not peremens.loaded_images_paths:
            messagebox.showwarning("Фу, как не культурно", "Пожалуйста, загрузите изображение.")
            log("Error, there is no image")
            return

        # if Functions.look():
        #     verify = messagebox.askyesno("Предупреждение!!!", "Программа не была должным образом протестированной,\n существующие таблицы могут нарушить работоспособность.\nВы хотите удальть старые таблицы?")
        #     if verify == True:
        #         Functions.clear_xlsm_down()
        #     else:
        #         verify = False
        #         messagebox.showwarning("НУ есть косяк, но времени нет", "Просим прощения")
        for i in range(len(peremens.loaded_images_paths)):
            peremens.filepath = peremens.loaded_images_paths[i]
            process_image(peremens.filepath)
            Functions.save_resultat()
            log(f"Processing {i+1} из {len(peremens.loaded_images_paths)}")
        Functions.save_resultat_global()
        log("Сompletion")
        messagebox.showinfo("Оно работает?", "Картинки обработались, файлы сохранены вместе с картинками")
        window.tabview.set("neiro")


# Создание окна Tkinter ---------------------------------------------------------------
root_window = ctk.CTk() # создаем окно
class window():

    def create_window(): # настройка окна
        # функция для создания окна
        root_window.title("NeiroBowlling") # название окна
        h = root_window.winfo_screenheight() # высота экрана рабочего стола
        w = root_window.winfo_screenwidth() # ширина экрана рабочего стола
        hr = 450# высота окна
        wr = 400# ширина окна
        root_window.geometry(f"{wr}x{hr}+{w//2-wr//2}+{h//2-hr//2}") # задаем размеры окна и отступ
        root_window.resizable(False, False) # разрешаем изменять размеры окна
        root_window.wm_iconbitmap() 
        root_window.iconphoto(False, ImageTk.PhotoImage(file=outside.path_icno)) 
    create_window()

    def create_logo(): # название и лого
        frame_logo = ctk.CTkFrame(root_window, fg_color="transparent")
        frame_logo.pack(side=tk.TOP, fill="x", expand=False, pady=20)
    
        frame_logo_na = ctk.CTkFrame(frame_logo, fg_color="transparent")
        frame_logo_na.pack(anchor=tk.CENTER)
    
        ctk.CTkLabel(frame_logo_na, text="NeiroBowlling", font=ctk.CTkFont("Helvetica", size=40, weight="bold")).pack(side=tk.LEFT,pady=0)# размещение

        logo = ctk.CTkImage(Image.open(outside.path_logo), size=(80, 80)) # загрузка логот
        ctk.CTkLabel(frame_logo_na, image=logo, text="").pack(side=tk.LEFT, pady=10) # размещение логотип
    create_logo()

    tabview = ctk.CTkTabview(master=root_window, corner_radius=25) # создание перекюлчателя страничек
    tabview.pack(side=tk.BOTTOM, fill="both", expand=True, padx=30, pady=30) # размещение переключателя

    def create_tab_single_mode(tabview): # создание странички для одиночных папок
        tabview.add("neiro") # добавление таба

        def upload_button(): # кнопка для выбора папки с файлами
            ctk.CTkButton(master=tabview.tab("neiro"), text="Выбрать папку", command=Functions.download_images, font=ctk.CTkFont("Helvetica", size=18, weight="bold"), corner_radius=20, height=40).pack(padx=5, pady=15, fill="x") # размещение кнопки для выбора папки c файлами  
        upload_button()

        def run_button(): # кнопка для запуска обработки
            ctk.CTkButton(master=tabview.tab("neiro"), text="Обработать Изображение", command=Functions.start, font=ctk.CTkFont("Helvetica", size=18, weight="bold"), corner_radius=20, height=40).pack(padx=5, pady=15, fill="x") # размещение кнопки для запуска
        run_button()
    create_tab_single_mode(tabview)

    # логирование в массив не ложить!!! я забил на него...
    tabview.add("devtool") # добавление таба
    dev_log = ctk.CTkTextbox(tabview.tab("devtool"), state="disable")  # изменено на "normal"
    def create_textbox(dev_log, tabview): #создание поля логирования
        # текстовое пространство для логирование
        dev_log.pack(padx=0, pady=0, fill="both")
        # создание для него scrollbar
        scrollbar = ctk.CTkScrollbar(tabview.tab("devtool"), command = dev_log.yview)
        scrollbar.pack(side=tk.RIGHT, fill="y")
 
        # связываем scrollbar с текстовым полем
        dev_log.configure(yscrollcommand=scrollbar.set)    
    create_textbox(dev_log, tabview)
log("Ready to work") 
root_window.mainloop() # запуск окна
